from collections import OrderedDict
from datetime import date
from tempfile import NamedTemporaryFile

from django.conf import settings
from django.db.models import Q, Sum
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ..models import (
    Availability, CorpContact, Corporation, Course, Section, Student, Teacher,
    Training,
)
from ..utils import school_year_start

openxml_contenttype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class OpenXMLExport:
    def __init__(self, sheet_title):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_title
        self.bold = Font(bold=True)
        self.row_idx = 1

    def write_line(self, values, bold=False, col_widths=()):
        for col_idx, value in enumerate(values, start=1):
            cell = self.ws.cell(row=self.row_idx, column=col_idx)
            try:
                cell.value = value
            except KeyError:
                # Ugly workaround for https://bugs.python.org/issue28969
                from openpyxl.utils.datetime import to_excel
                to_excel.cache_clear()
                cell.value = value
            if bold:
                cell.font = self.bold
            if col_widths:
                self.ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
        self.row_idx += 1

    def get_http_response(self, filename_base):
        with NamedTemporaryFile() as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            response = HttpResponse(tmp, content_type=openxml_contenttype)
            response['Content-Disposition'] = 'attachment; filename=%s_%s.xlsx' % (
                filename_base, date.strftime(date.today(), '%Y-%m-%d')
            )
        return response


EXPORT_FIELDS = [
    # Student fields
    ('ID externe', 'student__ext_id'),
    ('Prénom', 'student__first_name'), ('Nom', 'student__last_name'),
    ('Titre', 'student__gender'),
    ('Classe', 'student__klass__name'),
    ('Filière', 'student__klass__section__name'),
    ('Rue élève', 'student__street'),
    ('NPA_élève', 'student__pcode'),
    ('Localité élève', 'student__city'),
    ('Tél élève', 'student__tel'),
    ('Email élève', 'student__email'),
    ('Date de naissance', 'student__birth_date'),
    ('No AVS', 'student__avs'),
    # Stage fields
    ('Nom de la pratique professionnelle', 'availability__period__title'),
    ('Début', 'availability__period__start_date'), ('Fin', 'availability__period__end_date'),
    ('Remarques pratique professionnelle', 'comment'),
    ('Prénom référent', 'referent__first_name'), ('Nom référent', 'referent__last_name'),
    ('Courriel référent', 'referent__email'),
    ('Institution', 'availability__corporation__name'),
    ('ID externe Inst', 'availability__corporation__ext_id'),
    ('Rue Inst', 'availability__corporation__street'),
    ('NPA Inst', 'availability__corporation__pcode'),
    ('Ville Inst', 'availability__corporation__city'),
    ('Tél Inst', 'availability__corporation__tel'),
    ('Domaine', 'availability__domain__name'),
    ('Remarques Inst', 'availability__comment'),
    ('Civilité contact', 'availability__contact__civility'),
    ('Prénom contact', 'availability__contact__first_name'),
    ('Nom contact', 'availability__contact__last_name'),
    ('ID externe contact', 'availability__contact__ext_id'),
    ('Tél contact', 'availability__contact__tel'),
    ('Courriel contact', 'availability__contact__email'),
    ('Courriel contact - copie', None),
]


NON_ATTR_EXPORT_FIELDS = [
    ('Filière', 'period__section__name'),
    ('Nom de la pratique professionnelle', 'period__title'),
    ('Début', 'period__start_date'), ('Fin', 'period__end_date'),
    ('Institution', 'corporation__name'),
    ('Rue Inst', 'corporation__street'),
    ('NPA Inst', 'corporation__pcode'),
    ('Ville Inst', 'corporation__city'),
    ('Tél Inst', 'corporation__tel'),
    ('Domaine', 'domain__name'),
    ('Remarques Inst', 'comment'),
    ('Civilité contact', 'contact__civility'),
    ('Prénom contact', 'contact__first_name'),
    ('Nom contact', 'contact__last_name'),
    ('Tél contact', 'contact__tel'),
    ('Courriel contact', 'contact__email'),
    ('Courriel contact - copie', None),
]


def stages_export(request, scope=None):
    period_filter = request.GET.get('period')
    non_attributed = bool(int(request.GET.get('non_attr', 0)))

    export_fields = OrderedDict(EXPORT_FIELDS)
    contact_test_field = 'availability__contact__last_name'
    corp_name_field = 'availability__corporation__name'

    if period_filter:
        if non_attributed:
            # Export non attributed availabilities for a specific period
            query = Availability.objects.filter(period_id=period_filter, training__isnull=True)
            export_fields = OrderedDict(NON_ATTR_EXPORT_FIELDS)
            contact_test_field = 'contact__last_name'
            corp_name_field = 'corporation__name'
        else:
            # Export trainings for a specific period
            query = Training.objects.filter(availability__period_id=period_filter)
    else:
        if scope and scope == 'all':
            # Export all trainings in the database
            query = Training.objects.all()
        else:
            query = Training.objects.filter(availability__period__end_date__gt=school_year_start())

    # Prepare "default" contacts (when not defined on training)
    section_names = Section.objects.all().values_list('name', flat=True)
    default_contacts = dict(
        (c, {s: '' for s in section_names})
        for c in Corporation.objects.all().values_list('name', flat=True)
    )
    always_ccs = dict(
        (c, {s: [] for s in section_names})
        for c in Corporation.objects.all().values_list('name', flat=True)
    )
    for contact in CorpContact.objects.filter(corporation__isnull=False
            ).select_related('corporation'
            ).prefetch_related('sections').order_by('corporation'):
        for section in contact.sections.all():
            if not default_contacts[contact.corporation.name][section.name] or contact.is_main is True:
                default_contacts[contact.corporation.name][section.name] = contact
            if contact.always_cc:
                always_ccs[contact.corporation.name][section.name].append(contact)
        if contact.is_main:
            for sname in section_names:
                if not default_contacts[contact.corporation.name][sname]:
                    default_contacts[contact.corporation.name][sname] = contact

    export = OpenXMLExport('Pratiques professionnelles')
    export.write_line(export_fields.keys(), bold=True)  # Headers
    # Data
    query_keys = [f for f in export_fields.values() if f is not None]
    for line in query.values(*query_keys):
        values = []
        for field in query_keys:
            value = line[field]
            if 'gender' in field:
                value = {'F': 'Madame', 'M': 'Monsieur', '': ''}[value]
            values.append(value)
        if line[contact_test_field] is None:
            # Use default contact
            contact = default_contacts.get(line[corp_name_field], {}).get(line[export_fields['Filière']])
            if contact:
                values = values[:-6] + [
                    contact.civility, contact.first_name, contact.last_name, contact.ext_id,
                    contact.tel, contact.email
                ]
        if always_ccs[line[corp_name_field]].get(line[export_fields['Filière']]):
            values.append("; ".join(
                [c.email for c in always_ccs[line[corp_name_field]].get(line[export_fields['Filière']])]
            ))
        export.write_line(values)

    return export.get_http_response('pp_export')


def _ratio_Ede_Ase_Assc():
    # Spliting for unattribued periods
    tot_edeps = Course.objects.filter(imputation='EDEps').aggregate(Sum('period'))['period__sum'] or 0
    tot_edepe = Course.objects.filter(imputation='EDEpe').aggregate(Sum('period'))['period__sum'] or 0
    edepe_ratio = 1 if tot_edepe + tot_edeps == 0 else tot_edepe / (tot_edepe + tot_edeps)

    tot_asefe = Course.objects.filter(imputation='ASEFE').aggregate(Sum('period'))['period__sum'] or 0
    tot_mpts = Course.objects.filter(imputation='MPTS').aggregate(Sum('period'))['period__sum'] or 0
    asefe_ratio = 1 if tot_asefe + tot_mpts == 0 else tot_asefe / (tot_asefe + tot_mpts)

    tot_asscfe = Course.objects.filter(imputation='ASSCFE').aggregate(Sum('period'))['period__sum'] or 0
    tot_mps = Course.objects.filter(imputation='MPS').aggregate(Sum('period'))['period__sum'] or 0
    asscfe_ratio = 1 if tot_asscfe + tot_mps == 0 else tot_asscfe / (tot_asscfe + tot_mps)

    return {'edepe': edepe_ratio, 'asefe': asefe_ratio, 'asscfe': asscfe_ratio}


def imputations_export(request):
    IMPUTATIONS_EXPORT_FIELDS = [
        'Nom', 'Prénom', 'Report passé', 'Ens', 'Discipline',
        'Accomp.', 'Discipline', 'Total payé', 'Indice', 'Taux', 'Report futur',
        'ASA', 'ASSC', 'ASE', 'MPTS', 'MPS', 'EDEpe', 'EDEps', 'EDS', 'CAS_FPP'
    ]

    ratios = _ratio_Ede_Ase_Assc()

    export = OpenXMLExport('Imputations')
    export.write_line(IMPUTATIONS_EXPORT_FIELDS, bold=True)  # Headers

    for teacher in Teacher.objects.filter(archived=False):
        activities, imputations = teacher.calc_imputations(ratios)
        values = [
            teacher.last_name, teacher.first_name, teacher.previous_report,
            activities['tot_ens'], 'Ens. prof.', activities['tot_mandats'] + activities['tot_formation'],
            'Accompagnement', activities['tot_paye'], 'Charge globale',
            '{0:.2f}'.format(activities['tot_paye']/settings.GLOBAL_CHARGE_PERCENT),
            teacher.next_report,
        ]
        values.extend(imputations.values())
        export.write_line(values)

    return export.get_http_response('Imputations_export')


def export_sap(request):
    EXPORT_SAP_HEADERS = [
        'PERNR', 'PERNOM', 'DEGDA', 'ENDDA', 'ZNOM', 'ZUND',
        'ZACT', 'ZBRA', 'ZOTP', 'ZCCO', 'ZORD', 'ZTAUX',
    ]
    MAPPING_OTP = {
        'ASAFE': 'CIFO01.03.02.03.01.02 - ASA EE',
        'ASEFE': 'CIFO01.03.02.04.01.02 - CFC ASE EE',
        'ASSCFE': 'CIFO01.03.02.04.02.02 - CFC ASSC EE',
        'EDEpe': 'CIFO01.03.02.07.01.01 - EDE prat. prof. PT',
        'EDEps': 'CIFO01.03.02.07.02.01 - EDE stages PT',
        'EDS': 'CIFO01.03.02.07.03.02 - EDS EE',
        'CAS_FPP': 'CIFO01.03.02.01.03 - Mandats divers (CAS FPP)',
        'MPTS' : 'CIFO01.04.03.06.02.01 - MPTS ASE',
        'MPS': 'CIFO01.04.03.06.03.01 - MPS Santé',
    }

    ratios = _ratio_Ede_Ase_Assc()

    export = OpenXMLExport('Imputations')
    export.write_line(EXPORT_SAP_HEADERS, bold=True)  # Headers
    start_date = '20.08.2018'
    end_date = '19.08.2019'
    indice = 'charge globale'
    type_act = 'Ens. prof.'
    branche = 'Ens. prof.'
    centre_cout = ''
    stat = ''

    for teacher in Teacher.objects.filter(archived=False):
        activities, imputations = teacher.calc_imputations(ratios)
        for key in imputations:
            if imputations[key] > 0:
                values = [
                    teacher.ext_id, teacher.full_name, start_date, end_date, imputations[key], indice, type_act,
                    branche, MAPPING_OTP[key], centre_cout, stat,
                    round(imputations[key] / settings.GLOBAL_CHARGE_PERCENT, 2),
                ]
                export.write_line(values)

        # Previous report
        values = [
            teacher.ext_id, teacher.full_name, start_date, end_date, teacher.previous_report, indice, type_act,
            branche, 'Report précédent', centre_cout, stat,
            round(teacher.previous_report / settings.GLOBAL_CHARGE_PERCENT, 2),
        ]
        export.write_line(values)

        # Next report
        values = [
            teacher.ext_id, teacher.full_name, start_date, end_date, teacher.next_report, indice, type_act,
            branche, 'Report suivant', centre_cout, stat,
            round(teacher.next_report / settings.GLOBAL_CHARGE_PERCENT, 2),
        ]
        export.write_line(values)
    return export.get_http_response('Export_SAP')


GENERAL_EXPORT_FIELDS = [
    ('Num_Ele', 'ext_id'),
    ('Nom_Ele', 'last_name'),
    ('Prenom_Ele', 'first_name'),
    ('Genre_Ele', 'gender'),
    ('Rue_Ele', 'street'),
    ('NPA_Ele', 'pcode'),
    ('Ville_Ele', 'city'),
    ('DateNaissance_Ele', 'birth_date'),
    ('NOAVS_Ele', 'avs'),
    ('Canton_Ele', 'district'),
    ('Email_Ele', 'email'),
    ('Mobile_Ele', 'mobile'),
    ('Compte_RPN', 'login_rpn'),
    ('DispenseCG_Ele', 'dispense_ecg'),
    ('DispenseEPS_Ele', 'dispense_eps'),
    ('SoutienDYS_Ele', 'soutien_dys'),

    ('Classe_Ele', 'klass__name'),
    ('Filiere_Ele', 'klass__section__name'),
    ('MaitreDeClasseNom_Ele', 'klass__teacher__last_name'),
    ('MaitreDeClassePrenom_Ele', 'klass__teacher__first_name'),
    ('OptionASE_Ele', 'option_ase__name'),

    ('Num_Emp', 'corporation__ext_id'),
    ('Nom_Emp', 'corporation__name'),
    ('Rue_Emp', 'corporation__street'),
    ('NPA_Emp', 'corporation__pcode'),
    ('Ville_Emp', 'corporation__city'),
    ('Canton_Emp', 'corporation__district'),
    ('Secteur_Emp', 'corporation__sector'),
    ('Type_EMP', 'corporation__typ'),
    ('Tel_Emp', 'corporation__tel'),

    ('Num_Form', 'instructor__ext_id'),
    ('Titre_Form', 'instructor__civility'),
    ('Prenom_Form', 'instructor__first_name'),
    ('Nom_Form', 'instructor__last_name'),
    ('Tel_Form', 'instructor__tel'),
    ('Email_Form', 'instructor__email'),
    ('Num_Form2', 'instructor2__ext_id'),
    ('Titre_Form2', 'instructor2__civility'),
    ('Prenom_Form2', 'instructor2__first_name'),
    ('Nom_Form2', 'instructor2__last_name'),
    ('Tel_Form2', 'instructor2__tel'),
    ('Email_Form2', 'instructor2__email'),
    ('EmailCopie_Form', None),
]


def general_export(request):
    """
    Export all current students data
    """
    export_fields = OrderedDict(GENERAL_EXPORT_FIELDS)
    export = OpenXMLExport('Exportation')
    export.write_line(export_fields.keys(), bold=True)  # Headers
    # Data
    query_keys = [f for f in export_fields.values() if f is not None]
    query = Student.objects.filter(archived=False).order_by('klass__name', 'last_name', 'first_name')
    for line in query.values(*query_keys):
        values = []
        for field in query_keys:
            if field == 'gender':
                values.append(('Madame', 'Monsieur')[line[field] == 'M'])
            elif field in ('dispense_ecg', 'dispense_eps', 'soutien_dys'):
                values.append('Oui' if line[field] is True else '')
            else:
                values.append(line[field])
        export.write_line(values)

    return export.get_http_response('general_export')


ORTRA_EXPORT_FIELDS = [
    ('Num_Ele', 'ext_id'),
    ('Nom_Ele', 'last_name'),
    ('Prenom_Ele', 'first_name'),
    ('Genre_Ele', 'gender'),
    ('Rue_Ele', 'street'),
    ('NPA_Ele', 'pcode'),
    ('Ville_Ele', 'city'),
    ('DateNaissance_Ele', 'birth_date'),
    ('Email_Ele', 'email'),
    ('Mobile_Ele', 'mobile'),

    ('Classe_Ele', 'klass__name'),
    ('Filiere_Ele', 'klass__section__name'),
    ('MaitreDeClasseNom_Ele', 'klass__teacher__last_name'),
    ('MaitreDeClassePrenom_Ele', 'klass__teacher__first_name'),
    ('OptionASE_Ele', 'option_ase__name'),

    ('Num_Emp', 'corporation__ext_id'),
    ('Nom_Emp', 'corporation__name'),
    ('Rue_Emp', 'corporation__street'),
    ('NPA_Emp', 'corporation__pcode'),
    ('Ville_Emp', 'corporation__city'),
    ('Tel_Emp', 'corporation__tel'),

    ('Titre_Form', 'instructor__civility'),
    ('Prenom_Form', 'instructor__first_name'),
    ('Nom_Form', 'instructor__last_name'),
    ('Tel_Form', 'instructor__tel'),
    ('Email_Form', 'instructor__email'),
]


def ortra_export(request):
    """
    Export students data from sections ASAFE, ASEFE and ASSCFE
    """
    export_fields = OrderedDict(ORTRA_EXPORT_FIELDS)
    export = OpenXMLExport('Exportation')
    export.write_line(export_fields.keys(), bold=True)  # Headers
    # Data
    query_keys = [f for f in export_fields.values() if f is not None]
    query = Student.objects.filter(Q(klass__name__contains='ASAFE') |
                                   Q(klass__name__contains='ASEFE') |
                                   Q(klass__name__contains='ASSCFE'),
                                   archived=False).order_by('klass__name',
                                                            'last_name',
                                                            'first_name')

    for line in query.values(*query_keys):
        values = []
        for field in query_keys:
            if field == 'gender':
                values.append(('Madame', 'Monsieur')[line[field] == 'M'])
            else:
                values.append(line[field])
        export.write_line(values)

    return export.get_http_response('ortra_export')


def export_qualification(request, section='ede'):
    headers = [
        'Classe', 'Etudiant-e',
        'Référent pratique', 'Titre TD', 'Résumé TD', 'Ens. référent',
        'Mentor',
        'Session', 'Type', 'Exp_int.',
        'Expert ext. Civilité', 'Expert ext. Nom', 'Expert ext. Adresse', 'Expert ext. Localité',
        'Date', 'Salle', 'Note',
    ]

    export_name = 'Export_qualif_%s' % section.upper()
    export = OpenXMLExport(export_name)
    export.write_line(headers, bold=True)

    # Data
    empty_values = [''] * 7
    for student in Student.objects.filter(klass__name__startswith='3%s' % section.upper(), archived=False
            ).select_related('klass', 'referent', 'training_referent', 'mentor',
            ).prefetch_related('examination_set'
            ).order_by('klass__name', 'last_name'):
        stud_values = [
            student.klass.name,
            student.full_name,
            student.training_referent.full_name if student.training_referent else '',
            student.title,
            student.subject,
            student.referent.full_name if student.referent else '',
            student.mentor.full_name if student.mentor else '',
        ]
        lines_exported = 0
        for exam in student.examination_set.all():
            exam_values = [
                str(exam.session),
                exam.get_type_exam_display(),
                exam.internal_expert.full_name if exam.internal_expert else '',
                exam.external_expert.civility if exam.external_expert else '',
                exam.external_expert.full_name if exam.external_expert else '',
                exam.external_expert.street if exam.external_expert else '',
                exam.external_expert.pcode_city if exam.external_expert else '',
                exam.date_exam,
                exam.room,
                exam.mark,
            ]
            if lines_exported == 0:
                export.write_line(stud_values + exam_values)
            else:
                export.write_line(empty_values + exam_values)
            lines_exported += 1
        if lines_exported == 0:
            export.write_line(stud_values)

    return export.get_http_response(export_name)


def institutions_export(request):
    def format_value(val):
        return '' if val is None else str(val)

    fields = [
        (f.verbose_name, f.name)
        for f in Corporation._meta.get_fields() if hasattr(f, 'verbose_name') and f.name not in ('archived',)
    ]
    headers = [f[0] for f in fields]

    export = OpenXMLExport('Institutions')
    export.write_line(headers, bold=True)
    for corp in Corporation.objects.filter(archived=False).order_by('name'):
        values = [format_value(getattr(corp, f[1])) for f in fields]
        export.write_line(values)
    return export.get_http_response('Institutions')
